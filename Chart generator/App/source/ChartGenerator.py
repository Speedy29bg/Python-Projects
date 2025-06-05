import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
import csv
import os
import re
from matplotlib.figure import Figure
from typing import List, Dict, Tuple, Optional, Any
import itertools
import threading
import queue
from datetime import datetime
import traceback

###############################################################################
# ChartGenerator01.py - Laboratory Chart Generator
# 
# PROGRAM SEGMENTS OVERVIEW:
#
# 1. Import & Setup:
#    Imports required libraries for UI (tkinter), data processing (pandas),
#    file operations, charting (matplotlib, openpyxl), and utilities.
#
# 2. Application Documentation:
#    Doxygen-style documentation for the file and main class, describing
#    the application's purpose and functionality.
#
# 3. Main Class Definition:
#    LabChartGenerator class that serves as the primary container for the application.
#    Initializes application variables, settings, and UI components.
#
# 4. UI Construction:
#    Creates the complete user interface with multiple sections including:
#    - File selection controls
#    - Axis selection area (X, Primary Y, Secondary Y)
#    - Chart scaling and formatting options
#    - Preview area
#    - Output settings
#
# 5. Chart Export Functionality:
#    Handles saving the currently displayed chart to different formats (Excel, PNG, PDF).
#
# 6. UI Control Functions:
#    Functions that manage UI interactions like enabling/disabling form elements
#    and file selection.
#
# 7. Data Loading & Processing:
#    Core functions for:
#    - CSV file loading with automatic format detection
#    - Header row identification
#    - Data preview
#    - Column type detection and conversion
#
# 8. Chart Generation:
#    Functions for creating and updating charts including:
#    - Data scaling and normalization
#    - Datetime handling
#    - Primary and secondary axis setup
#    - Plot customization
#    - Chart preview generation
#
# 9. Excel Chart Export:
#    Functions to generate multi-sheet Excel workbooks containing:
#    - Data from selected columns
#    - Line or scatter charts
#    - Progress tracking and error handling
#
# 10. Main Program Entry Point:
#     Creates and starts the application when the script is run directly.
###############################################################################

## 
# @file ChartGenerator01.py
# @brief A comprehensive chart generation tool for laboratory data analysis
#
# @details This application provides a graphical interface for generating and customizing charts
# from CSV data files. It supports line and scatter plots, multiple Y-axes, log scaling, 
# data normalization, and export to various formats. The tool is particularly optimized for
# working with laboratory time-series data that might contain datetime information.
#
# @author Lab Chart Tools Team
# @version 1.0
# @date April 2025

##
# @class LabChartGenerator
# @brief Main class for the Laboratory Chart Generator application
#
# @details This class provides a comprehensive GUI for loading, visualizing and exporting
# chart data from CSV files. It supports various chart formatting options, including
# multiple axes, logarithmic scaling, and normalization. 
# The application can detect datetime columns and properly format them on chart axes.
class LabChartGenerator:
    ##
    # @brief Constructor initializes the application window and variables
    #
    # @param root The tkinter root window instance
    #
    # @details Creates a new instance of the LabChartGenerator application with all
    # necessary variables initialized, including file lists, data caches, chart settings, etc.
    def __init__(self, root):
        self.root = root
        self.root.title("Lab Chart Generator")
        self.root.geometry("1300x1000")
        
        # Set application icon if available
        try:
            self.root.iconbitmap("chart_icon.ico")
        except:
            pass  # Icon file not found, use default
        
        # Variables
        self.files: List[str] = []
        self.canvas = None
        self.figure = None
        self.column_headers: List[str] = []
        self.x_selected_indices: List[int] = []
        self.y_selected_indices: List[int] = []
        self.secondary_selected_indices: List[int] = []
        self.file_data_cache: Dict[str, pd.DataFrame] = {}  # Cache for loaded file data
        
        # Chart settings
        self.auto_scale = tk.BooleanVar(value=True)
        self.log_scale_x = tk.BooleanVar(value=False)
        self.log_scale_y1 = tk.BooleanVar(value=False)
        self.log_scale_y2 = tk.BooleanVar(value=False)
        self.normalize_data = tk.BooleanVar(value=False)
        
        # Custom range variables
        self.x_min = tk.StringVar()
        self.x_max = tk.StringVar()
        self.y1_min = tk.StringVar()
        self.y1_max = tk.StringVar()
        self.y2_min = tk.StringVar()
        self.y2_max = tk.StringVar()
        
        # Chart type and export options
        self.chart_type = tk.StringVar(value="line")
        self.export_format = tk.StringVar(value="xlsx")
        
        # Color scheme for charts
        self.color_schemes = ["default", "viridis", "plasma", "inferno", "magma", "cividis"]
        self.selected_color_scheme = tk.StringVar(value=self.color_schemes[0])
        
        # Setup logging
        self.setup_logging()
        
        # Setup UI
        self.setup_ui()
    
    ##
    # @brief Sets up logging for the application
    #
    # @details Configures logging to both console and file, with different levels
    # to help with troubleshooting and debugging application issues.
    def setup_logging(self):
        """Configure application logging"""
        import logging
        
        # Create logs directory if it doesn't exist
        logs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
        os.makedirs(logs_dir, exist_ok=True)
        
        # Configure root logger
        logger = logging.getLogger()
        logger.setLevel(logging.DEBUG)  # Capture all levels
        
        # Clear existing handlers to avoid duplicates
        for handler in logger.handlers[:]:
            logger.removeHandler(handler)
        
        # Create formatters
        file_formatter = logging.Formatter(
            '%(asctime)s [%(levelname)s] %(filename)s:%(lineno)d - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        console_formatter = logging.Formatter('%(levelname)s: %(message)s')
        
        # File handler - detailed logs
        log_filename = f"chart_generator_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        file_handler = logging.FileHandler(os.path.join(logs_dir, log_filename))
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(file_formatter)
        logger.addHandler(file_handler)
        
        # Console handler - less detailed for console
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)  # Less verbose for console
        console_handler.setFormatter(console_formatter)
        logger.addHandler(console_handler)
        
        logging.info("Logging system initialized")
        logging.info(f"Log file: {os.path.join(logs_dir, log_filename)}")
        
        # Store reference to logger
        self.logger = logger
        
    ##
    # @brief Exception handler for the application
    #
    # @param exc_type The type of exception
    # @param exc_value The exception instance
    # @param exc_traceback The traceback object
    #
    # @details This function catches unhandled exceptions, logs them with
    # full traceback, and displays a user-friendly error message.
    def handle_exception(self, exc_type, exc_value, exc_traceback):
        """Handle uncaught exceptions"""
        # Log the error with full traceback
        import logging
        logging.error("Uncaught exception", 
                     exc_info=(exc_type, exc_value, exc_traceback))
        
        # Format a user-friendly error message
        error_msg = f"An unexpected error occurred:\n{exc_value}"
        
        # Show error dialog if tkinter is still functioning
        try:
            import tkinter.messagebox as messagebox
            messagebox.showerror("Application Error", error_msg)
        except:
            # If tkinter is not working, print to console
            print(error_msg)
        
        # Update status label if possible
        try:
            if hasattr(self, 'status_label'):
                self.status_label.config(text=f"Error: {str(exc_value)}")
        except:
            pass

    ##
    # @brief Creates and arranges all UI elements in the application window
    #
    # @details Sets up the complete interface including file selection area, axis selection
    # listboxes, chart scaling options, preview area, and control buttons. The layout uses a
    # combination of frames and layout managers to organize the interface into logical sections.
    def setup_ui(self):
        """Create and arrange UI elements"""
        # Main frame
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # File selection
        file_frame = ttk.LabelFrame(main_frame, text="File Selection")
        file_frame.pack(fill=tk.X, pady=5)
        
        self.file_label = ttk.Label(file_frame, text="No files selected")
        self.file_label.pack(side=tk.LEFT, padx=10, pady=5)
        
        ttk.Button(file_frame, text="Browse", command=self.select_files).pack(side=tk.RIGHT, padx=10, pady=5)
        ttk.Button(file_frame, text="Preview Data", command=self.preview_data).pack(side=tk.RIGHT, padx=10, pady=5)
        
        # Axes selection frame
        axes_frame = ttk.LabelFrame(main_frame, text="Axes Selection")
        axes_frame.pack(fill=tk.X, pady=5)
        
        # Create three columns for axes selection
        axes_left = ttk.Frame(axes_frame)
        axes_middle = ttk.Frame(axes_frame)
        axes_right = ttk.Frame(axes_frame)
        
        axes_left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        axes_middle.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        axes_right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # X-Axis selection
        ttk.Label(axes_left, text="X-Axis (Ctrl+click):").pack(anchor=tk.W)
        x_frame = ttk.Frame(axes_left)
        x_frame.pack(fill=tk.BOTH, expand=True)
        
        self.x_listbox = tk.Listbox(x_frame, selectmode="multiple", height=6, width=30, exportselection=0)
        x_scrollbar = ttk.Scrollbar(x_frame, orient="vertical", command=self.x_listbox.yview)
        
        self.x_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        x_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.x_listbox.config(yscrollcommand=x_scrollbar.set)
        self.x_listbox.bind("<<ListboxSelect>>", lambda e: self.update_preview())
        
        # Primary Y-Axes selection
        ttk.Label(axes_middle, text="Primary Y-Axes (Ctrl+click):").pack(anchor=tk.W)
        y_frame = ttk.Frame(axes_middle)
        y_frame.pack(fill=tk.BOTH, expand=True)
        
        self.y_listbox = tk.Listbox(y_frame, selectmode="multiple", height=6, width=30, exportselection=0)
        y_scrollbar = ttk.Scrollbar(y_frame, orient="vertical", command=self.y_listbox.yview)
        
        self.y_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.y_listbox.config(yscrollcommand=y_scrollbar.set)
        self.y_listbox.bind("<<ListboxSelect>>", lambda e: self.update_preview())
        
        # Secondary Y-Axes selection
        ttk.Label(axes_right, text="Secondary Y-Axes (Ctrl+click):").pack(anchor=tk.W)
        secondary_frame = ttk.Frame(axes_right)
        secondary_frame.pack(fill=tk.BOTH, expand=True)
        
        self.secondary_listbox = tk.Listbox(secondary_frame, selectmode="multiple", height=6, width=30, exportselection=0)
        secondary_scrollbar = ttk.Scrollbar(secondary_frame, orient="vertical", command=self.secondary_listbox.yview)
        
        self.secondary_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        secondary_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.secondary_listbox.config(yscrollcommand=secondary_scrollbar.set)
        self.secondary_listbox.bind("<<ListboxSelect>>", lambda e: self.update_preview())
        
        # Chart scaling options
        scaling_frame = ttk.LabelFrame(main_frame, text="Chart Scaling Options")
        scaling_frame.pack(fill=tk.X, pady=5)
        
        # Left and right sections for scaling options
        scale_left = ttk.Frame(scaling_frame)
        scale_right = ttk.Frame(scaling_frame)
        scale_left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scale_right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Scale options
        ttk.Checkbutton(scale_left, text="Auto Scale", variable=self.auto_scale, 
                      command=self.toggle_manual_scaling).pack(anchor=tk.W)
        ttk.Checkbutton(scale_left, text="Log Scale X-Axis", variable=self.log_scale_x,
                      command=self.update_preview).pack(anchor=tk.W)
        ttk.Checkbutton(scale_left, text="Log Scale Primary Y-Axis", variable=self.log_scale_y1,
                      command=self.update_preview).pack(anchor=tk.W)
        ttk.Checkbutton(scale_left, text="Log Scale Secondary Y-Axis", variable=self.log_scale_y2,
                      command=self.update_preview).pack(anchor=tk.W)
        ttk.Checkbutton(scale_left, text="Normalize Data (0-1)", variable=self.normalize_data,
                      command=self.update_preview).pack(anchor=tk.W)
        
        # Chart type selection
        ttk.Label(scale_left, text="Chart Type:").pack(anchor=tk.W, pady=(10, 0))
        chart_type_frame = ttk.Frame(scale_left)
        chart_type_frame.pack(anchor=tk.W)
        
        ttk.Radiobutton(chart_type_frame, text="Line", variable=self.chart_type, 
                      value="line", command=self.update_preview).pack(side=tk.LEFT)
        ttk.Radiobutton(chart_type_frame, text="Scatter", variable=self.chart_type, 
                      value="scatter", command=self.update_preview).pack(side=tk.LEFT)
        
        # Export format selection
        ttk.Label(scale_left, text="Export Format:").pack(anchor=tk.W, pady=(10, 0))
        export_format_frame = ttk.Frame(scale_left)
        export_format_frame.pack(anchor=tk.W)
        
        ttk.Radiobutton(export_format_frame, text="Excel", variable=self.export_format, 
                      value="xlsx").pack(side=tk.LEFT)
        ttk.Radiobutton(export_format_frame, text="PNG", variable=self.export_format, 
                      value="png").pack(side=tk.LEFT)
        ttk.Radiobutton(export_format_frame, text="PDF", variable=self.export_format, 
                      value="pdf").pack(side=tk.LEFT)
        
        # Custom range inputs
        range_frame = ttk.Frame(scale_right)
        range_frame.pack(fill=tk.X)
        
        # X-Axis range
        ttk.Label(range_frame, text="X-Axis Range:").grid(row=0, column=0, sticky=tk.W, pady=2)
        ttk.Label(range_frame, text="Min:").grid(row=0, column=1, sticky=tk.E, padx=5, pady=2)
        self.x_min_entry = ttk.Entry(range_frame, textvariable=self.x_min, width=10, state='disabled')
        self.x_min_entry.grid(row=0, column=2, pady=2)
        ttk.Label(range_frame, text="Max:").grid(row=0, column=3, sticky=tk.E, padx=5, pady=2)
        self.x_max_entry = ttk.Entry(range_frame, textvariable=self.x_max, width=10, state='disabled')
        self.x_max_entry.grid(row=0, column=4, pady=2)
        
        # Primary Y-Axis range
        ttk.Label(range_frame, text="Primary Y-Axis Range:").grid(row=1, column=0, sticky=tk.W, pady=2)
        ttk.Label(range_frame, text="Min:").grid(row=1, column=1, sticky=tk.E, padx=5, pady=2)
        self.y1_min_entry = ttk.Entry(range_frame, textvariable=self.y1_min, width=10, state='disabled')
        self.y1_min_entry.grid(row=1, column=2, pady=2)
        ttk.Label(range_frame, text="Max:").grid(row=1, column=3, sticky=tk.E, padx=5, pady=2)
        self.y1_max_entry = ttk.Entry(range_frame, textvariable=self.y1_max, width=10, state='disabled')
        self.y1_max_entry.grid(row=1, column=4, pady=2)
        
        # Secondary Y-Axis range
        ttk.Label(range_frame, text="Secondary Y-Axis Range:").grid(row=2, column=0, sticky=tk.W, pady=2)
        ttk.Label(range_frame, text="Min:").grid(row=2, column=1, sticky=tk.E, padx=5, pady=2)
        self.y2_min_entry = ttk.Entry(range_frame, textvariable=self.y2_min, width=10, state='disabled')
        self.y2_min_entry.grid(row=2, column=2, pady=2)
        ttk.Label(range_frame, text="Max:").grid(row=2, column=3, sticky=tk.E, padx=5, pady=2)
        self.y2_max_entry = ttk.Entry(range_frame, textvariable=self.y2_max, width=10, state='disabled')
        self.y2_max_entry.grid(row=2, column=4, pady=2)
        
        # Store references to the entry widgets for enabling/disabling
        self.range_entries = [
            self.x_min_entry, self.x_max_entry, 
            self.y1_min_entry, self.y1_max_entry, 
            self.y2_min_entry, self.y2_max_entry
        ]
        
        # Color scheme selection
        color_scheme_frame = ttk.LabelFrame(main_frame, text="Color Scheme")
        color_scheme_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(color_scheme_frame, text="Select Color Scheme:").pack(side=tk.LEFT, padx=5, pady=5)
        color_scheme_menu = ttk.OptionMenu(color_scheme_frame, self.selected_color_scheme, *self.color_schemes)
        color_scheme_menu.pack(side=tk.LEFT, padx=5, pady=5)

        # Preview frame
        preview_frame = ttk.LabelFrame(main_frame, text="Chart Preview")
        preview_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.preview_frame = ttk.Frame(preview_frame)
        self.preview_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Output options
        output_frame = ttk.LabelFrame(main_frame, text="Output Options")
        output_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(output_frame, text="Output Filename:").pack(side=tk.LEFT, padx=5, pady=5)
        self.output_filename = ttk.Entry(output_frame, width=30)
        self.output_filename.insert(0, "Lab_Charts")
        self.output_filename.pack(side=tk.LEFT, padx=5, pady=5)
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        #TODO - change the place of the button for gentae charts to a different button frame
        ttk.Button(scale_left, text="Generate Charts", command=self.generate_charts).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear Selections", command=self.clear_selections).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Apply Custom Scaling", command=self.update_preview).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Export Current Chart", command=self.export_current_chart).pack(side=tk.LEFT, padx=5)
        
        # Status bar
        self.status_label = ttk.Label(main_frame, text="Ready", relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.pack(fill=tk.X, side=tk.BOTTOM, pady=5)
    
    ##
    # @brief Exports the currently displayed chart to the selected format
    #
    # @details Saves the current matplotlib figure to a file in the specified format (Excel, PNG or PDF).
    # The user is prompted to choose a save location through a file dialog.
    #
    # @exception Exception Raised if the export process fails
    def export_current_chart(self):
        """Export the currently displayed chart to the selected format"""
        if not self.figure:
            messagebox.showinfo("Info", "No chart to export")
            return
        
        try:
            # Get output filename without extension
            base_filename = self.output_filename.get()
            if not base_filename:
                base_filename = "Chart"
            
            # Strip any extension
            base_filename = os.path.splitext(base_filename)[0]
            
            # Add appropriate extension based on selected format
            format_type = self.export_format.get()
            filename = f"{base_filename}.{format_type}"
            
            # Ask for save location
            save_path = filedialog.asksaveasfilename(
                defaultextension=f".{format_type}",
                filetypes=[(f"{format_type.upper()} files", f"*.{format_type}")],
                initialfile=filename
            )
            
            if not save_path:
                return  # User cancelled
            
            # Save the figure
            self.figure.savefig(save_path, format=format_type, dpi=300, bbox_inches='tight')
            
            self.status_label.config(text=f"Chart exported as {os.path.basename(save_path)}")
            messagebox.showinfo("Success", f"Chart exported successfully to {save_path}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export chart: {str(e)}")
            self.status_label.config(text=f"Error exporting chart: {str(e)}")

    ## 
    # @brief Enables or disables manual scaling input fields
    #
    # @details Changes the state of all range entry widgets based on the auto_scale checkbox.
    # When auto scaling is enabled, the manual range fields are disabled.
    def toggle_manual_scaling(self):
        """Enable or disable manual scaling inputs"""
        state = 'disabled' if self.auto_scale.get() else 'normal'
        
        # Update state for all range entry widgets
        for entry in self.range_entries:
            entry.configure(state=state)
        
        self.update_preview()
    
    ## 
    # @brief Opens a file dialog for CSV file selection
    #
    # @details Allows the user to select one or multiple CSV files for analysis. Updates
    # the file list, clears the file cache for files no longer selected, and loads headers
    # from the first selected file into the column selection listboxes.
    #
    # @exception Exception Raised if file loading fails
    def select_files(self):
        """Handle file selection and load column headers"""
        selected_files = filedialog.askopenfilenames(filetypes=[("CSV files", "*.csv")])
        
        if not selected_files:
            return
        
        self.files = selected_files
        self.file_label.config(text=f"Selected: {len(self.files)} files")
        
        try:
            # Clear cache for files that are no longer selected
            self.file_data_cache = {k: v for k, v in self.file_data_cache.items() if k in self.files}
            
            # Load files in background
            self.load_files_in_background(self.files, self.detect_headers_and_populate_listboxes)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file(s): {str(e)}")
            self.status_label.config(text="Error loading files")
    
    ## 
    # @brief Loads files in a background thread to keep UI responsive
    #
    # @param files List of file paths to load
    # @param callback Function to call when loading is complete
    #
    # @details Creates a background thread to load multiple CSV files without
    # freezing the UI. Updates the status label during loading and calls the
    # provided callback function when complete.
    def load_files_in_background(self, files, callback):
        """Load files in background thread to keep UI responsive"""
        self.status_label.config(text="Loading files...")
        
        # Create loading indicator
        loading_frame = ttk.Frame(self.root)
        loading_frame.place(relx=0.5, rely=0.5, anchor='center')
        
        loading_label = ttk.Label(loading_frame, text="Loading files...", font=('Helvetica', 12, 'bold'))
        loading_label.pack(pady=10)
        
        progress = ttk.Progressbar(loading_frame, orient="horizontal", length=300, mode="indeterminate")
        progress.pack(pady=10)
        progress.start(10)
        
        # Create queue for thread communication
        file_queue = queue.Queue()
        
        # Function to run in background thread
        def load_files_thread():
            loaded_data = {}
            
            for i, file_path in enumerate(files):
                try:
                    header_row_idx = self.detect_header_row(file_path)
                    df = self.read_csv_file(file_path, header_row_idx)
                    loaded_data[file_path] = df
                    
                    # Update progress info via queue
                    file_queue.put(("progress", i+1, len(files), os.path.basename(file_path)))
                except Exception as e:
                    file_queue.put(("error", os.path.basename(file_path), str(e)))
            
            # Signal completion
            file_queue.put(("done", loaded_data))
        
        # Function to process queue updates in main thread
        def check_queue():
            try:
                while True:
                    message = file_queue.get_nowait()
                    
                    if message[0] == "progress":
                        _, current, total, filename = message
                        loading_label.config(text=f"Loading {current}/{total}: {filename}")
                    
                    elif message[0] == "error":
                        _, filename, error = message
                        print(f"Error loading {filename}: {error}")
                    
                    elif message[0] == "done":
                        _, loaded_data = message
                        # Update cache with loaded data
                        self.file_data_cache.update(loaded_data)
                        
                        # Remove loading indicator
                        loading_frame.destroy()
                        
                        # Call the callback
                        if callback:
                            callback()
                        
                        return  # Exit the update loop
            
            except queue.Empty:
                # Queue is empty but loading not complete, check again after a delay
                self.root.after(100, check_queue)
        
        # Start background thread
        threading.Thread(target=load_files_thread, daemon=True).start()
        
        # Start queue checking in main thread
        check_queue()
    
    ## 
    # @brief Shows a preview of the data in the first selected file
    #
    # @details Opens a new window displaying the first 100 rows of the selected CSV file
    # in a scrollable text widget, allowing the user to inspect the raw data.
    #
    # @exception Exception Raised if data preview fails
    def preview_data(self):
        """Show a preview of the data in the first selected file"""
        if not self.files:
            messagebox.showinfo("Info", "No files selected")
            return
        
        try:
            # Get data from the first file
            file_path = self.files[0]
            if file_path in self.file_data_cache:
                df = self.file_data_cache[file_path]
            else:
                header_row_idx = self.detect_header_row(file_path)
                df = self.read_csv_file(file_path, header_row_idx)
                self.file_data_cache[file_path] = df
            
            if df.empty:
                messagebox.showinfo("Info", "No data found in the file")
                return
            
            # Create a new window for the preview
            preview_window = tk.Toplevel(self.root)
            preview_window.title(f"Data Preview: {os.path.basename(file_path)}")
            preview_window.geometry("800x600")
            
            # Create a Text widget to display the data
            text_widget = tk.Text(preview_window, wrap="none")
            
            # Add scrollbars
            y_scrollbar = ttk.Scrollbar(preview_window, orient="vertical", command=text_widget.yview)
            x_scrollbar = ttk.Scrollbar(preview_window, orient="horizontal", command=text_widget.xview)
            text_widget.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
            
            # Layout
            text_widget.grid(row=0, column=0, sticky="nsew")
            y_scrollbar.grid(row=0, column=1, sticky="ns")
            x_scrollbar.grid(row=1, column=0, sticky="ew")
            
            # Configure grid weights
            preview_window.grid_rowconfigure(0, weight=1)
            preview_window.grid_columnconfigure(0, weight=1)
            
            # Display dataframe as formatted text
            text_widget.insert(tk.END, df.head(100).to_string())
            text_widget.configure(state="disabled")  # Make read-only
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to preview data: {str(e)}")
            self.status_label.config(text="Error previewing data")
    
    ## 
    # @brief Detects headers from the first file and populates column selection listboxes
    #
    # @details Reads headers from the first selected CSV file, populates all three listboxes 
    # (X-axis, Primary Y-axis, Secondary Y-axis) with these headers, and resets selection indices.
    #
    # @exception Exception Raised if header detection fails
    def detect_headers_and_populate_listboxes(self):
        """Detect headers from the first file and populate listboxes"""
        if not self.files:
            return
        
        header_row_idx = self.detect_header_row(self.files[0])
        
        try:
            # Check if the file data is already cached
            if self.files[0] in self.file_data_cache:
                df = self.file_data_cache[self.files[0]]
            else:
                df = self.read_csv_file(self.files[0], header_row_idx)
                self.file_data_cache[self.files[0]] = df
            
            self.column_headers = df.columns.tolist()
            
            # Clear and populate listboxes
            self.x_listbox.delete(0, tk.END)
            self.y_listbox.delete(0, tk.END)
            self.secondary_listbox.delete(0, tk.END)
            
            if self.column_headers and len(self.column_headers) > 0:
                for header in self.column_headers:
                    self.x_listbox.insert(tk.END, header)
                    self.y_listbox.insert(tk.END, header)
                    self.secondary_listbox.insert(tk.END, header)
            else:
                self.x_listbox.insert(tk.END, "No headers found")
                self.y_listbox.insert(tk.END, "No headers found")
                self.secondary_listbox.insert(tk.END, "No headers found")
            
            # Reset selection indices
            self.x_selected_indices = []
            self.y_selected_indices = []
            self.secondary_selected_indices = []
            
        except Exception as e:
            raise Exception(f"Error detecting headers: {str(e)}")
    
    ## 
    # @brief Analyzes the file to determine the row containing column headers
    #
    # @param file_path The absolute path to the CSV file
    # @return int The index of the header row (0-based)
    #
    # @details Uses multiple strategies to detect the header row:
    # 1. Looks for rows with mostly non-numeric content
    # 2. Falls back to the first non-empty row
    # 3. Uses first row (index 0) as last resort
    def detect_header_row(self, file_path):
        """Detect the row containing headers"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                reader = csv.reader(f)
                rows = list(itertools.islice(reader, 0, 20))  # Read first 20 rows for analysis
                
                # Check if file is empty
                if not rows:
                    print("File is empty")
                    return 0
                
                # Strategy 1: Look for the first row where most cells contain non-numeric content
                # This often indicates headers
                for i, row in enumerate(rows):
                    if len(row) >= 4:  # Need at least 2 columns to be useful
                        # Check if this row has mostly text (non-numeric) content
                        non_numeric_count = sum(1 for cell in row if cell.strip() and not cell.replace('.', '', 1).isdigit())
                        if non_numeric_count > len(row) / 2:  # More than half are non-numeric
                            print(f"Detected header at row {i} (mostly text content)")
                            return i
                
                # Strategy 2: Fall back to first non-empty row
                for i, row in enumerate(rows):
                    if any(cell.strip() for cell in row):
                        print(f"Using first non-empty row at index {i}")
                        return i
            
            # If all strategies fail, use the first row
            print("Using first row as header by default")
            return 0
        except Exception as e:
            print(f"Error detecting header row: {str(e)}")
            return 0
    
    ## 
    # @brief Reads a CSV file with automatic encoding and delimiter detection
    #
    # @param file_path The absolute path to the CSV file
    # @param header_row_idx The index of the header row (0-based)
    # @return pd.DataFrame A pandas DataFrame containing the CSV data
    #
    # @details Uses multiple strategies to read CSV files with potential encoding or delimiter issues:
    # 1. Attempts to detect the delimiter using csv.Sniffer
    # 2. Tries to read with pandas using UTF-8 encoding
    # 3. Falls back to manual CSV reading if pandas fails
    # 4. Tries alternative encodings (latin1, cp1252, iso-8859-1)
    #
    # @exception pd.errors.EmptyDataError Raised if the file is empty
    # @exception Exception Raised if all reading methods fail
    def read_csv_file(self, file_path, header_row_idx):
        """Read CSV file with appropriate encoding and delimiter detection"""
        try:
            print(f"Attempting to read file: {file_path}")
            
            # Try to detect delimiter
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                sample = f.read(4096)  # Read a sample of the file
            
            # Detect delimiter
            delimiter = ','  # Default delimiter
            try:
                sniffer = csv.Sniffer()
                if sample.strip():  # Make sure the sample isn't empty
                    detected_delimiter = sniffer.sniff(sample).delimiter
                    print(f"Detected delimiter: '{detected_delimiter}'")
                    delimiter = detected_delimiter
            except Exception as e:
                print(f"Delimiter detection failed: {str(e)}. Using default delimiter: ','")
            
            # First try with pandas - the standard approach
            try:
                print(f"Reading CSV with pandas, header_row_idx: {header_row_idx}, delimiter: '{delimiter}'")
                # Direct approach - read the file and specify the header row
                df = pd.read_csv(file_path, encoding='utf-8', header=header_row_idx, delimiter=delimiter, 
                                low_memory=False, on_bad_lines='warn')
                
                # Check if we have data
                if not df.empty:
                    print(f"CSV loaded successfully with pandas: {len(df)} rows, {len(df.columns)} columns")
                    print(f"Column headers: {list(df.columns)}")
                    return df
                else:
                    print("Warning: DataFrame is empty after reading with pandas")
            except Exception as e:
                print(f"Error reading with pandas: {str(e)}")
            
            # Fall back to manual CSV reading if pandas fails
            try:
                print("Falling back to manual CSV reading...")
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    csv_reader = csv.reader(f, delimiter=delimiter)
                    rows = list(csv_reader)
                
                if not rows:
                    print("File appears to be empty")
                    return pd.DataFrame()
                
                if header_row_idx >= len(rows):
                    print(f"Header row index ({header_row_idx}) out of range. Using first row.")
                    header_row_idx = 0
                
                # Extract headers and data
                headers = rows[header_row_idx]
                data = rows[header_row_idx+1:] if header_row_idx+1 < len(rows) else []
                
                # Create DataFrame
                df = pd.DataFrame(data, columns=headers)
                
                # Convert numeric columns
                for col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='ignore')
                
                print(f"Manual CSV reading successful: {len(df)} rows, {len(df.columns)} columns")
                return df
                
            except Exception as e:
                print(f"Manual CSV reading failed: {str(e)}")
                
            # Try alternative encodings if UTF-8 fails
            for encoding in ['latin1', 'cp1252', 'iso-8859-1']:
                try:
                    print(f"Trying with encoding: {encoding}")
                    df = pd.read_csv(file_path, encoding=encoding, header=header_row_idx, 
                                    delimiter=delimiter, low_memory=False)
                    print(f"Successful load with {encoding} encoding")
                    return df
                except Exception as e:
                    print(f"Failed with {encoding} encoding: {str(e)}")
            
            # As a last resort, try to read without specifying encoding
            try:
                print("Last attempt: reading without specifying encoding")
                df = pd.read_csv(file_path, header=header_row_idx, delimiter=delimiter, 
                                low_memory=False, encoding_errors='replace')
                return df
            except Exception as e:
                print(f"Final attempt failed: {str(e)}")
                
            # If all attempts fail
            raise Exception("All methods to read the CSV file failed")
            
        except pd.errors.EmptyDataError:
            print(f"Warning: File {os.path.basename(file_path)} is empty.")
            messagebox.showwarning("Warning", f"File {os.path.basename(file_path)} is empty.")
            return pd.DataFrame()
        except Exception as e:
            print(f"Error reading CSV file: {str(e)}")
            raise Exception(f"Failed to read CSV file: {str(e)}")
    
    ## 
    # @brief Updates the chart preview based on current selections
    #
    # @details Generates a preview chart based on the current axis selections, scaling options,
    # and display settings. Reads data from the first selected file, processes it according
    # to scaling options, and creates a new preview plot.
    #
    # @exception Exception Raised if chart generation fails
    def update_preview(self):
        """Update chart preview based on current selections"""
        if not self.files or "No headers found" in self.column_headers:
            return
        
        # Store current selections
        self.x_selected_indices = list(self.x_listbox.curselection())
        self.y_selected_indices = list(self.y_listbox.curselection())
        self.secondary_selected_indices = list(self.secondary_listbox.curselection())
        
        if not self.x_selected_indices or not self.y_selected_indices:
            self.status_label.config(text="Please select at least one X-axis and one Y-axis")
            return
        
        # Get selected items
        x_selected = [self.x_listbox.get(i) for i in self.x_selected_indices]
        y_selected = [self.y_listbox.get(i) for i in self.y_selected_indices]
        secondary_selected = [self.secondary_listbox.get(i) for i in self.secondary_selected_indices]
        
        try:
            # Read data from the first file (use cache if available)
            file_path = self.files[0]
            if file_path in self.file_data_cache:
                df = self.file_data_cache[file_path]
            else:
                header_row_idx = self.detect_header_row(file_path)
                df = self.read_csv_file(file_path, header_row_idx)
                self.file_data_cache[file_path] = df
            
            if df.empty:
                self.status_label.config(text="No data found in the file")
                return
            
            # Check if selected columns exist in dataframe
            all_selected = x_selected + y_selected + secondary_selected
            missing_columns = [col for col in all_selected if col not in df.columns]
            
            if missing_columns:
                self.status_label.config(text=f"Missing columns in data: {', '.join(missing_columns)}")
                return
            
            # Clean existing canvas
            self.clear_canvas()
            
            # Process data for scaling
            processed_df = self.process_data_for_scaling(df, x_selected[0], y_selected, secondary_selected)
            
            # Create new figure and plot
            self.create_preview_plot(processed_df, x_selected[0], y_selected, secondary_selected)
            
            self.status_label.config(text="Preview updated")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update preview: {str(e)}")
            self.status_label.config(text=f"Error: {str(e)}")
    
    ## 
    # @brief Processes DataFrame to implement scaling and transformation options
    #
    # @param df The input pandas DataFrame
    # @param x_axis The name of the X-axis column
    # @param y_axes List of primary Y-axis column names
    # @param secondary_axes List of secondary Y-axis column names
    # @return pd.DataFrame A processed copy of the input DataFrame
    #
    # @details Applies several data transformations based on user settings:
    # 1. Detects and converts datetime columns to numeric timestamps
    # 2. Converts non-numeric columns to numeric where possible
    # 3. Handles log scaling by replacing zeros/negative values
    # 4. Applies normalization to scale data to 0-1 range if enabled
    def process_data_for_scaling(self, df, x_axis, y_axes, secondary_axes):
        """Process dataframe for scaling options with improved error handling"""
        # Make a copy to avoid modifying original
        processed_df = df.copy()
        
        # Convert columns to numeric when possible
        columns_to_process = [x_axis] + y_axes + secondary_axes
        
        for col in columns_to_process:
            if col in processed_df.columns:  # Check if column exists
                # First, try to detect if this is a datetime column
                is_datetime = False
                
                # Check a sample of values for datetime format
                sample = processed_df[col].dropna().head(20).astype(str)
                date_patterns = [
                    # Common datetime patterns
                    r'\d{1,2}/\d{1,2}/\d{2,4}\s+\d{1,2}:\d{1,2}:\d{1,2}\s*(?:AM|PM)?',  # 1/16/2024 4:44:01 PM
                    r'\d{1,2}-\d{1,2}-\d{2,4}\s+\d{1,2}:\d{1,2}:\d{1,2}',  # 16-01-2024 16:44:01
                    r'\d{4}-\d{1,2}-\d{1,2}\s+\d{1,2}:\d{1,2}:\d{1,2}',  # 2024-01-16 16:44:01
                    r'\d{1,2}/\d{1,2}/\d{2,4}'  # 1/16/2024
                ]
                
                # Check if most values match a datetime pattern
                for pattern in date_patterns:
                    matches = sample.str.match(pattern).sum()
                    if matches > len(sample) * 0.5:  # If more than 50% match
                        is_datetime = True
                        print(f"Column '{col}' detected as datetime")
                        break
                
                if is_datetime:
                    try:
                        # Convert to datetime and then to timestamp for numeric processing
                        processed_df[col] = pd.to_datetime(processed_df[col], errors='coerce')
                        # Convert datetime to numeric (Unix timestamp in seconds)
                        processed_df[col] = processed_df[col].astype('int64') // 10**9
                        print(f"Converted datetime column '{col}' to timestamp values")
                    except Exception as e:
                        print(f"Warning: Failed to convert datetime column '{col}': {str(e)}")
                else:
                    # Try to convert to numeric if not already
                    if not pd.api.types.is_numeric_dtype(processed_df[col]):
                        try:
                            processed_df[col] = pd.to_numeric(processed_df[col], errors='coerce')
                            # Log the conversion
                            print(f"Converted column '{col}' to numeric with NaN for non-numeric values")
                        except Exception as e:
                            print(f"Warning: Could not convert column '{col}' to numeric: {str(e)}")
            else:
                print(f"Warning: Column '{col}' does not exist in the dataframe")
        
        # For log scales, replace zeros and negative values with small positive values
        if self.log_scale_x.get() and x_axis in processed_df.columns:
            # Replace negative/zero values for log scale
            min_positive = processed_df[x_axis][processed_df[x_axis] > 0].min() if any(processed_df[x_axis] > 0) else 0.001
            epsilon = min_positive / 10 if min_positive > 0 else 0.001
            processed_df[x_axis] = processed_df[x_axis].apply(lambda x: max(x, epsilon) if pd.notnull(x) else x)
        
        # Handle log scale for primary Y axes
        if self.log_scale_y1.get():
            for y_axis in y_axes:
                if y_axis in processed_df.columns:
                    min_positive = processed_df[y_axis][processed_df[y_axis] > 0].min() if any(processed_df[y_axis] > 0) else 0.001
                    epsilon = min_positive / 10 if min_positive > 0 else 0.001
                    processed_df[y_axis] = processed_df[y_axis].apply(lambda x: max(x, epsilon) if pd.notnull(x) else x)
        
        # Handle log scale for secondary Y axes
        if self.log_scale_y2.get():
            for sec_axis in secondary_axes:
                if sec_axis in processed_df.columns:
                    min_positive = processed_df[sec_axis][processed_df[sec_axis] > 0].min() if any(processed_df[sec_axis] > 0) else 0.001
                    epsilon = min_positive / 10 if min_positive > 0 else 0.001
                    processed_df[sec_axis] = processed_df[sec_axis].apply(lambda x: max(x, epsilon) if pd.notnull(x) else x)
        
        # Normalize data if selected
        if self.normalize_data.get():
            for col in columns_to_process:
                if col in processed_df.columns and pd.api.types.is_numeric_dtype(processed_df[col]):
                    col_min = processed_df[col].min()
                    col_max = processed_df[col].max()
                    if col_max > col_min:  # Prevent division by zero
                        processed_df[col] = (processed_df[col] - col_min) / (col_max - col_min)
        
        return processed_df
    
    ## 
    # @brief Clears the existing chart canvas
    #
    # @details Removes the current matplotlib canvas from the UI, closes the 
    # figure to free up resources, and resets the figure and canvas variables.
    def clear_canvas(self):
        """Clear the existing canvas"""
        if self.canvas:
            self.canvas.get_tk_widget().pack_forget()
            plt.close(self.figure)
            self.figure = None
            self.canvas = None
    
    ## 
    # @brief Creates and displays a preview chart based on the current data and settings
    #
    # @param df The processed pandas DataFrame containing the data to plot
    # @param x_axis The name of the X-axis column
    # @param y_axes List of primary Y-axis column names
    # @param secondary_axes List of secondary Y-axis column names
    #
    # @details This comprehensive function:
    # 1. Creates a matplotlib figure and axes
    # 2. Detects if X-axis is a datetime and formats it appropriately
    # 3. Applies selected scaling options (log, manual ranges)
    # 4. Creates line or scatter plots for primary and secondary axes
    # 5. Handles legend creation and plot formatting
    # 6. Embeds the plot in the application window
    def create_preview_plot(self, df, x_axis, y_axes, secondary_axes):
        """Create and display the preview plot"""
        # First, ensure we have data to plot
        if df.empty:
            messagebox.showinfo("Warning", "No data available for plotting")
            self.status_label.config(text="No data available for plotting")
            return
            
        # Debug information - print dataframe info to console
        print(f"DataFrame info for plotting:")
        print(f"- Shape: {df.shape}")
        print(f"- Columns: {df.columns.tolist()}")
        print(f"- Data types: {df.dtypes}")
        print(f"- X-axis column '{x_axis}' has {df[x_axis].count()} non-null values")
        for y_col in y_axes + secondary_axes:
            print(f"- Column '{y_col}' has {df[y_col].count()} non-null values")
        
        # Check if X-axis might be a datetime column in the original data
        x_is_datetime = False
        original_df = None
        original_datetime_values = None
        
        for file_path in self.files:
            if file_path in self.file_data_cache:
                original_df = self.file_data_cache[file_path]
                break
        
        if original_df is not None and x_axis in original_df.columns:
            # Check a sample of values for datetime format
            sample = original_df[x_axis].dropna().head(20).astype(str)
            date_patterns = [
                r'\d{1,2}/\d{1,2}/\d{2,4}\s+\d{1,2}:\d{1,2}:\d{1,2}\s*(?:AM|PM)?',  # 1/16/2024 4:44:01 PM
                r'\d{1,2}-\d{1,2}-\d{2,4}\s+\d{1,2}:\d{1,2}:\d{1,2}',  # 16-01-2024 16:44:01
                r'\d{4}-\d{1,2}-\d{1,2}\s+\d{1,2}:\d{1,2}:\d{1,2}',  # 2024-01-16 16:44:01
                r'\d{1,2}/\d{1,2}/\d{2,4}'  # 1/16/2024
            ]
            
            for pattern in date_patterns:
                matches = sample.str.match(pattern).sum()
                if matches > len(sample) * 0.5:  # If more than 50% match
                    x_is_datetime = True
                    print(f"X-axis '{x_axis}' detected as datetime")
                    
                    # Convert original values to datetime objects
                    try:
                        original_datetime_values = pd.to_datetime(original_df[x_axis], errors='coerce')
                        print(f"Successfully converted '{x_axis}' to datetime objects")
                    except Exception as e:
                        print(f"Error converting to datetime: {str(e)}")
                    
                    break
        
        # Create a copy of the dataframe for plotting
        plot_df = df.copy()
        
        # If x-axis is datetime, create a special dataframe with string timestamps for display
        if x_is_datetime and original_datetime_values is not None:
            # Create a version of the dataframe where the x-axis is a string representation
            # of the datetime with 15-minute rounding
            try:
                # Round timestamps to nearest 15 minutes for display
                rounded_times = original_datetime_values.dt.floor('15min')
                
                # Create string formatted dates for x-axis values
                time_strings = rounded_times.dt.strftime('%m/%d/%Y %H:%M')
                
                # Create a mapping from original timestamps to rounded string values
                # First convert original datetimes to Unix timestamps (same as in process_data_for_scaling)
                unix_timestamps = original_datetime_values.astype('int64') // 10**9
                
                # Create a mapping dictionary
                time_map = dict(zip(unix_timestamps, time_strings))
                
                # Store the mapping for future reference
                self.datetime_string_map = time_map
                
                print(f"Created datetime string mapping with {len(time_map)} entries")
            except Exception as e:
                print(f"Error creating datetime string mapping: {str(e)}")
        
        # Ensure data is numeric for plotting
        try:
            # Convert X-axis
            if not pd.api.types.is_numeric_dtype(plot_df[x_axis]):
                plot_df[x_axis] = pd.to_numeric(plot_df[x_axis], errors='coerce')
                print(f"Converted '{x_axis}' to numeric: {plot_df[x_axis].count()} valid values")
                
            # Convert Y-axes columns
            for col in y_axes + secondary_axes:
                if not pd.api.types.is_numeric_dtype(plot_df[col]):
                    plot_df[col] = pd.to_numeric(plot_df[col], errors='coerce')
                    print(f"Converted '{col}' to numeric: {plot_df[col].count()} valid values")
        except Exception as e:
            print(f"Error during numeric conversion: {str(e)}")
        
        # Remove rows with NaN in the x-axis
        plot_df = plot_df.dropna(subset=[x_axis])
        if plot_df.empty:
            messagebox.showinfo("Warning", f"No valid numeric data for X-axis '{x_axis}'")
            self.status_label.config(text=f"No valid numeric data for X-axis '{x_axis}'")
            return

        # Create figure
        self.figure, ax = plt.subplots(figsize=(10, 6))
        
        # Set log scales if selected
        if self.log_scale_x.get():
            ax.set_xscale('log')
        if self.log_scale_y1.get():
            ax.set_yscale('log')
        
        # Set manual axis limits if not auto-scaling
        if not self.auto_scale.get():
            try:
                # X-axis limits
                if self.x_min.get() and self.x_max.get():
                    ax.set_xlim(float(self.x_min.get()), float(self.x_max.get()))
                
                # Primary Y-axis limits
                if self.y1_min.get() and self.y1_max.get():
                    ax.set_ylim(float(self.y1_min.get()), float(self.y1_max.get()))
            except ValueError:
                messagebox.showwarning("Warning", "Invalid numeric values in range fields")
        
        # Plot based on chart type
        is_scatter = self.chart_type.get() == "scatter"
        has_valid_data = False
        
        # Plot primary Y-axes
        for y_axis in y_axes:
            # Get valid data points for this series (both x and y must be valid)
            series_df = plot_df[[x_axis, y_axis]].dropna()
            if len(series_df) == 0:
                print(f"No valid data points for {y_axis} vs {x_axis}")
                continue
                
            print(f"Plotting {y_axis} vs {x_axis}: {len(series_df)} points")
            has_valid_data = True
            
            if is_scatter:
                ax.scatter(series_df[x_axis], series_df[y_axis], label=y_axis, alpha=0.7)
            else:
                ax.plot(series_df[x_axis], series_df[y_axis], label=y_axis)
        
        # Set labels
        ax.set_xlabel(x_axis)
        y1_label = "Primary Y-Axes" if len(y_axes) > 1 else y_axes[0] if y_axes else ""
        if self.normalize_data.get() and y1_label:
            y1_label += " (Normalized)"
        ax.set_ylabel(y1_label)
        
        # Add grid
        ax.grid(True, linestyle='--', alpha=0.7)
        
        # Plot secondary Y-axes if selected
        if secondary_axes:
            ax2 = ax.twinx()
            
            # Set log scale for secondary Y-axis if selected
            if self.log_scale_y2.get():
                ax2.set_yscale('log')
            
            # Set manual limits for secondary Y-axis if not auto-scaling
            if not self.auto_scale.get():
                try:
                    if self.y2_min.get() and self.y2_max.get():
                        ax2.set_ylim(float(self.y2_min.get()), float(self.y2_max.get()))
                except ValueError:
                    pass  # Already warned above
            
            # Plot secondary axes
            for sec_axis in secondary_axes:
                # Get valid data points for this series
                series_df = plot_df[[x_axis, sec_axis]].dropna()
                if len(series_df) == 0:
                    print(f"No valid data points for {sec_axis} vs {x_axis}")
                    continue
                    
                print(f"Plotting {sec_axis} vs {x_axis}: {len(series_df)} points")
                has_valid_data = True
                
                if is_scatter:
                    ax2.scatter(series_df[x_axis], series_df[sec_axis], label=sec_axis, marker='x', alpha=0.7)
                else:
                    ax2.plot(series_df[x_axis], series_df[sec_axis], label=sec_axis, linestyle="--")
            
            # Set label for secondary Y-axis
            y2_label = "Secondary Y-Axes" if len(secondary_axes) > 1 else secondary_axes[0] if secondary_axes else ""
            if self.normalize_data.get() and y2_label:
                y2_label += " (Normalized)"
            ax2.set_ylabel(y2_label)
            
            # Add legend for secondary Y-axis
            lines2, labels2 = ax2.get_legend_handles_labels()
            if lines2:
                ax2.legend(lines2, labels2, loc="upper right")
        
        # Add legend for primary axes
        lines1, labels1 = ax.get_legend_handles_labels()
        if lines1:
            ax.legend(lines1, labels1, loc="upper left")
        
        # Check if we actually plotted anything
        if not has_valid_data:
            plt.close(self.figure)
            self.figure = None
            messagebox.showinfo("Warning", "No valid data points to plot")
            self.status_label.config(text="No valid data points to plot")
            return
        
        # Format X-axis ticks for datetime if needed
        if x_is_datetime:
            try:
                # Import necessary modules
                import numpy as np
                
                # Get current x-axis tick positions
                locs = ax.get_xticks()
                
                if hasattr(self, 'datetime_string_map') and self.datetime_string_map:
                    # Create a custom formatter for the x-axis
                    def format_timestamp(x, pos=None):
                        # Find the nearest mapped timestamp
                        if not np.isfinite(x):
                            return ''
                        
                        # Round to nearest integer as the map uses integers
                        x_int = int(round(x))
                        
                        # Find the closest key in the map
                        closest_keys = sorted(self.datetime_string_map.keys(), 
                                              key=lambda k: abs(k - x_int))
                        
                        if closest_keys:
                            # Get the datetime string from the map
                            return self.datetime_string_map[closest_keys[0]]
                        else:
                            return str(x_int)
                    
                    # Apply the custom formatter
                    from matplotlib.ticker import FuncFormatter
                    ax.xaxis.set_major_formatter(FuncFormatter(format_timestamp))
                    
                    # Set tick positions at regular intervals
                    if locs.size > 0:
                        # Get the min and max of data
                        x_min, x_max = plot_df[x_axis].min(), plot_df[x_axis].max()
                        
                        # Create about 5-10 ticks across the range
                        total_range = x_max - x_min
                        tick_interval = max(1, total_range // 8)  # At least 1 unit apart
                        
                        # Generate tick positions
                        tick_positions = np.arange(
                            np.floor(x_min / tick_interval) * tick_interval,
                            np.ceil(x_max / tick_interval) * tick_interval + tick_interval,
                            tick_interval
                        )
                        
                        if len(tick_positions) > 1:
                            ax.set_xticks(tick_positions)
                
                # Rotate the tick labels for better readability
                plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
                
                print("Successfully formatted datetime axis with string labels")
            except Exception as e:
                print(f"Warning: Error formatting datetime axis: {str(e)}")
                # Continue with standard numeric axis as fallback
        
        # Set title
        plot_type = "Scatter" if is_scatter else "Line"
        scaling_info = " (Normalized)" if self.normalize_data.get() else ""
        ax.set_title(f"{plot_type} Chart: {', '.join(y_axes)} vs {x_axis}{scaling_info}")
        
        # Adjust layout to make room for labels
        plt.tight_layout()
        
        # Create canvas
        self.canvas = FigureCanvasTkAgg(self.figure, master=self.preview_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        self.status_label.config(text="Chart preview updated successfully")
    
    ## 
    # @brief Clears all axis selections and preview
    #
    # @details Deselects all items in all three listboxes, resets selection indices, 
    # clears the canvas, and updates the status label.
    def clear_selections(self):
        """Clear all selections"""
        self.x_listbox.selection_clear(0, tk.END)
        self.y_listbox.selection_clear(0, tk.END)
        self.secondary_listbox.selection_clear(0, tk.END)
        self.x_selected_indices = []
        self.y_selected_indices = []
        self.secondary_selected_indices = []
        self.clear_canvas()
        self.status_label.config(text="Selections cleared")
    
    ## 
    # @brief Generates charts for all selected files and exports to Excel
    #
    # @details Creates a multi-sheet Excel workbook with:
    # 1. One sheet per input CSV file
    # 2. Data from selected columns copied to each sheet
    # 3. Line or scatter charts embedded in each sheet
    # 4. All currently selected formatting options applied
    #
    # Shows a progress dialog during generation and handles cancellation.
    #
    # @exception Exception Raised if chart generation or Excel export fails
    def generate_charts(self):
        """Generate charts for all selected files"""
        if not self.files:
            messagebox.showinfo("Info", "No files selected")
            return
        
        x_selected = [self.x_listbox.get(i) for i in self.x_listbox.curselection()]
        y_selected = [self.y_listbox.get(i) for i in self.y_listbox.curselection()]
        secondary_selected = [self.secondary_listbox.get(i) for i in self.secondary_listbox.curselection()]
        
        if not x_selected or not y_selected:
            messagebox.showinfo("Info", "Please select at least one X-axis and one Y-axis")
            return
        
        x_axis = x_selected[0]
        
        try:
            output_filename = self.output_filename.get()
            if not output_filename.endswith('.xlsx'):
                output_filename += '.xlsx'
            
            # Create progress dialog
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Generating Charts")
            progress_window.geometry("300x150")
            progress_window.transient(self.root)
            progress_window.grab_set()
            
            # Add a label and progress bar
            ttk.Label(progress_window, text="Processing files...").pack(pady=10)
            progress_var = tk.DoubleVar()
            progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=len(self.files))
            progress_bar.pack(fill=tk.X, padx=20, pady=10)
            file_label = ttk.Label(progress_window, text="")
            file_label.pack(pady=10)
            
            # Add a cancel button
            cancel_button = ttk.Button(progress_window, text="Cancel", command=progress_window.destroy)
            cancel_button.pack(pady=10)
            
            # Initialize workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            processed_files = 0
            skipped_files = 0
            
            for file_idx, file in enumerate(self.files):
                try:
                    # Update progress
                    progress_var.set(file_idx + 1)
                    file_label.config(text=f"Processing: {os.path.basename(file)}")
                    progress_window.update()
                    
                    # Check if user cancelled
                    if not progress_window.winfo_exists():
                        return
                    
                    # Create safe sheet name (Excel has 31 char limit)
                    base_name = os.path.basename(file).replace('.csv', '')
                    safe_sheet_name = self.create_safe_sheet_name(base_name)
                    
                    # Read data
                    header_row_idx = self.detect_header_row(file)
                    df = self.read_csv_file(file, header_row_idx)
                    
                    # Check if dataframe is empty
                    if df.empty:
                        skipped_files += 1
                        continue
                    
                    # Check if all selected columns exist in this file
                    all_selected = [x_axis] + y_selected + secondary_selected
                    missing_columns = [col for col in all_selected if col not in df.columns]
                    
                    if missing_columns:
                        file_label.config(text=f"Skipping {base_name}: Missing columns")
                        progress_window.update()
                        skipped_files += 1
                        continue
                    
                    # Process data for scaling if needed
                    if self.normalize_data.get() or self.log_scale_x.get() or self.log_scale_y1.get() or self.log_scale_y2.get():
                        df = self.process_data_for_scaling(df, x_axis, y_selected, secondary_selected)
                    
                    # Create sheet and populate data
                    ws = wb.create_sheet(safe_sheet_name)
                    self.populate_worksheet(ws, df, x_axis, y_selected, secondary_selected)
                    
                    # Create chart based on selected type
                    if self.chart_type.get() == "scatter":
                        self.add_scatter_chart_to_worksheet(ws, df, x_axis, y_selected, secondary_selected)
                    else:
                        self.add_line_chart_to_worksheet(ws, df, x_axis, y_selected, secondary_selected)

                    processed_files += 1
                except Exception as e:
                    file_label.config(text=f"Error processing {os.path.basename(file)}")
                    progress_window.update()
                    skipped_files += 1
                    print(f"Error processing {file}: {str(e)}")
            
            # Close progress window
            progress_window.destroy()
            
            # Save workbook
            wb.save(output_filename)
            
            # Show success message
            message = f"Charts generated successfully!\n\n" \
                      f"- Files processed: {processed_files}\n" \
                      f"- Files skipped: {skipped_files}\n" \
                      f"- Output file: {output_filename}"
            
            messagebox.showinfo("Success", message)
            self.status_label.config(text=f"Charts saved as {output_filename}")
            
        except Exception as e:
            try:
                progress_window.destroy()
            except:
                pass
            messagebox.showerror("Error", f"Failed to generate charts: {str(e)}")
            self.status_label.config(text=f"Error: {str(e)}")

    ## 
    # @brief Creates a safe sheet name for Excel by removing invalid characters
    #
    # @param base_name The original file name to convert to sheet name
    # @return str A sanitized sheet name compatible with Excel requirements
    #
    # @details Removes special characters that are not allowed in Excel sheet names
    # and truncates the name to 31 characters (Excel's limit).
    def create_safe_sheet_name(self, base_name):
        """Create a safe sheet name by removing invalid characters and truncating"""
        safe_name = re.sub(r'[\\/*?:"<>|]', '_', base_name)
        return safe_name[:31]

    ## 
    # @brief Populates an Excel worksheet with data from a DataFrame
    #
    # @param ws The openpyxl worksheet object
    # @param df The pandas DataFrame containing source data
    # @param x_axis The name of the X-axis column
    # @param y_axes List of primary Y-axis column names
    # @param secondary_axes List of secondary Y-axis column names
    #
    # @details Writes column headers in row 1, then writes all data values
    # starting from row 2, handling NaN values appropriately.
    def populate_worksheet(self, ws, df, x_axis, y_axes, secondary_axes):
        """Populate worksheet with data"""
        all_axes = [x_axis] + y_axes + secondary_axes
        
        # Write headers
        for col_idx, axis in enumerate(all_axes, start=1):
            ws.cell(row=1, column=col_idx, value=axis)
        
        # Write data - use numeric index to avoid issues with DataFrame index
        for idx, (_, row) in enumerate(df.iterrows(), start=0):
            for col_idx, axis in enumerate(all_axes, start=1):
                # Handle NaN values
                value = row[axis]
                if pd.isna(value):
                    value = None
                ws.cell(row=idx+2, column=col_idx, value=value)

    ## 
    # @brief Adds a line chart to an Excel worksheet
    #
    # @param ws The openpyxl worksheet object
    # @param df The pandas DataFrame containing source data
    # @param x_axis The name of the X-axis column
    # @param y_axes List of primary Y-axis column names
    # @param secondary_axes List of secondary Y-axis column names
    #
    # @details Creates an openpyxl LineChart object with:
    # 1. Data references to the worksheet cells
    # 2. Primary and secondary Y-axes properly configured
    # 3. Titles and labels set appropriately
    # The chart is inserted into the worksheet at position D5.
    def add_line_chart_to_worksheet(self, ws, df, x_axis, y_axes, secondary_axes):
        """Add a line chart to the worksheet"""
        chart = LineChart()
        chart.title = f"Line Chart: {', '.join(y_axes)} vs {x_axis}"
        chart.x_axis.title = x_axis
        chart.y_axis.title = "Primary Y-Axes"
        
        # Add primary Y-axis data
        for idx, y_axis in enumerate(y_axes, start=2):
            data = Reference(ws, min_col=idx, min_row=1, max_row=len(df)+1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
        
        # Add secondary Y-axis data if available
        if secondary_axes:
            # Create a second y-axis for the chart
            chart.y_axis.crosses = "min"
            chart.y_axis.axId = 200
            
            second_y_axis = chart.y_axis
            second_y_axis.title = "Secondary Y-Axes"
            second_y_axis.axId = 201
            second_y_axis.crosses = "max"
            
            # Add secondary series with different style
            for idx, sec_axis in enumerate(secondary_axes, start=len(y_axes)+2):
                data = Reference(ws, min_col=idx, min_row=1, max_row=len(df)+1)
                s = chart.add_data(data, titles_from_data=True)
                # Mark this series to use the secondary axis
                s.axId = second_y_axis.axId
        
        ws.add_chart(chart, "D5")

    ## 
    # @brief Adds a scatter chart to an Excel worksheet
    #
    # @param ws The openpyxl worksheet object
    # @param df The pandas DataFrame containing source data
    # @param x_axis The name of the X-axis column
    # @param y_axes List of primary Y-axis column names
    # @param secondary_axes List of secondary Y-axis column names
    #
    # @details Creates an openpyxl ScatterChart object with:
    # 1. X and Y data references to the worksheet cells
    # 2. Primary and secondary Y-axes properly configured 
    # 3. Titles and labels set appropriately
    # The chart is inserted into the worksheet at position D5.
    def add_scatter_chart_to_worksheet(self, ws, df, x_axis, y_axes, secondary_axes):
        """Add a scatter chart to the worksheet"""
        chart = ScatterChart()
        chart.title = f"Scatter Chart: {', '.join(y_axes)} vs {x_axis}"
        chart.x_axis.title = x_axis
        chart.y_axis.title = "Primary Y-Axes"
        
        # Add primary Y-axis series
        x_values = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
        
        for idx, y_axis in enumerate(y_axes, start=2):
            y_values = Reference(ws, min_col=idx, min_row=1, max_row=len(df)+1)
            series = chart.series.append(y_values, x_values)
            series.title = Reference(ws, min_col=idx, min_row=1, max_row=1)
        
        # Add secondary Y-axis series if available
        if secondary_axes:
            # Create a second y-axis
            chart.y_axis.crosses = "min"
            chart.y_axis.axId = 200
            
            second_y_axis = chart.y_axis
            second_y_axis.title = "Secondary Y-Axes"
            second_y_axis.axId = 201
            second_y_axis.crosses = "max"
            
            for idx, sec_axis in enumerate(secondary_axes, start=len(y_axes)+2):
                y_values = Reference(ws, min_col=idx, min_row=1, max_row=len(df)+1)
                series = chart.series.append(y_values, x_values)
                series.title = Reference(ws, min_col=idx, min_row=1, max_row=1)
                # Mark this series to use the secondary axis
                series.axId = second_y_axis.axId
        
        ws.add_chart(chart, "D5")

if __name__ == "__main__":
    root = tk.Tk()
    app = LabChartGenerator(root)
    # Set custom exception handler
    import sys
    sys.excepthook = app.handle_exception
    root.mainloop()
    root.destroy()
