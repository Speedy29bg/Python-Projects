"""
Excel Export Module

This module provides functions for generating Excel workbooks with chart data.

Author: Lab Chart Tools Team
"""

from openpyxl import Workbook
import pandas as pd
import logging
import os
from typing import List, Dict, Any

# Import necessary functions from other modules
from modules.export_module.worksheet_utils import populate_worksheet, add_line_chart_to_worksheet
from modules.data_processor.header_detection import detect_header_row
from modules.data_processor.csv_reader import read_csv_file
from modules.data_processor.data_scaling import process_data_for_scaling
from modules.data_processor.utils import create_safe_sheet_name

def generate_excel_workbook(files: List[str], file_data_cache: Dict[str, pd.DataFrame], 
                           x_axis: str, y_axes: List[str], secondary_axes: List[str],
                           scaling_options: Dict[str, Any], output_name: str) -> Dict[str, Any]:
    """Generate an Excel workbook with charts from multiple data files
    
    Args:
        files: List of file paths to process
        file_data_cache: Dictionary mapping file paths to DataFrames
        x_axis: The name of the X-axis column
        y_axes: List of primary Y-axis column names
        secondary_axes: List of secondary Y-axis column names
        scaling_options: Dictionary of scaling options
        output_name: Base name for the output file
        
    Returns:
        Dictionary with success status and output filename or error message
    """
    result = {
        'success': False,
        'filename': '',
        'error': ''
    }
    
    if not files:
        result['error'] = "No files provided"
        return result
        
    try:
        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
            
        # Get chart type from scaling options
        chart_type = scaling_options.get('chart_type', 'line')
        
        # Track processed files
        processed_files = 0
        skipped_files = 0
        
        for file_idx, file_path in enumerate(files):
            try:
                logging.info(f"Processing file {file_idx + 1}/{len(files)}: {os.path.basename(file_path)}")
                
                # Get data from cache or load if not present
                if file_path in file_data_cache:
                    df = file_data_cache[file_path]
                else:
                    header_row_idx = detect_header_row(file_path)
                    df = read_csv_file(file_path, header_row_idx)
                    file_data_cache[file_path] = df
                
                if df.empty:
                    logging.warning(f"Skipping empty file: {os.path.basename(file_path)}")
                    skipped_files += 1
                    continue
                    
                # Check if selected columns exist in the dataframe
                all_columns = [x_axis] + y_axes + secondary_axes
                missing_columns = [col for col in all_columns if col not in df.columns]
                
                if missing_columns:
                    logging.warning(f"Skipping file {os.path.basename(file_path)} due to missing columns: {', '.join(missing_columns)}")
                    skipped_files += 1
                    continue
                    
                # Process data for scaling
                processed_df = process_data_for_scaling(df, x_axis, y_axes, secondary_axes, scaling_options)
                
                # Create sheet name from filename
                sheet_name = create_safe_sheet_name(os.path.basename(file_path))
                if sheet_name in wb.sheetnames:
                    sheet_name = f"{sheet_name}_{file_idx + 1}"
                    
                # Create new worksheet
                ws = wb.create_sheet(title=sheet_name)
                
                # Populate data
                populate_worksheet(ws, processed_df, x_axis, y_axes, secondary_axes)
                
                # Add chart based on selected type
                if chart_type == 'scatter':
                    add_scatter_chart_to_worksheet(ws, processed_df, x_axis, y_axes, secondary_axes)
                else:  # Default to line chart
                    add_line_chart_to_worksheet(ws, processed_df, x_axis, y_axes, secondary_axes)
                    
                processed_files += 1
                
            except Exception as e:
                logging.error(f"Error processing file {os.path.basename(file_path)}: {str(e)}")
                skipped_files += 1
        
        # Set result based on processed files
        if processed_files > 0:
            # Save the workbook
            filename = f"{output_name}.xlsx"
            wb.save(filename)
            
            result['success'] = True
            result['filename'] = filename
            logging.info(f"Successfully processed {processed_files} files, skipped {skipped_files} files")
        else:
            result['error'] = f"No files could be processed. {skipped_files} files were skipped due to errors."
            logging.warning(result['error'])
            
    except Exception as e:
        result['error'] = str(e)
        logging.error(f"Error generating Excel workbook: {str(e)}")
    
    return result
