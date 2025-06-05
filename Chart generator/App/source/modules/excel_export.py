from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
import pandas as pd
import logging
from typing import List, Dict, Tuple, Optional, Any
import os

def populate_worksheet(ws, df, x_axis, y_axes, secondary_axes):
    """Populate worksheet with data
    
    Args:
        ws: The openpyxl worksheet object
        df: The pandas DataFrame containing source data
        x_axis: The name of the X-axis column
        y_axes: List of primary Y-axis column names
        secondary_axes: List of secondary Y-axis column names
    """
    all_axes = [x_axis] + y_axes + secondary_axes
    
    # Write headers
    for col_idx, axis in enumerate(all_axes, start=1):
        ws.cell(row=1, column=col_idx, value=axis)
    
    # Write data - use numeric index to avoid issues with DataFrame index
    for idx, (_, row) in enumerate(df.iterrows(), start=0):
        for col_idx, axis in enumerate(all_axes, start=1):
            # Handle NaN values
            value = row[axis]
            if pd.isna(value):
                value = None
            ws.cell(row=idx+2, column=col_idx, value=value)
    
    logging.debug(f"Populated worksheet with {len(df)} rows and {len(all_axes)} columns")

def add_line_chart_to_worksheet(ws, df, x_axis, y_axes, secondary_axes):
    """Add a line chart to the worksheet
    
    Args:
        ws: The openpyxl worksheet object
        df: The pandas DataFrame containing source data
        x_axis: The name of the X-axis column
        y_axes: List of primary Y-axis column names
        secondary_axes: List of secondary Y-axis column names
    """
    chart = LineChart()
    chart.title = f"Line Chart: {', '.join(y_axes)} vs {x_axis}"
    chart.x_axis.title = x_axis
    chart.y_axis.title = "Primary Y-Axes"
    
    # Add primary Y-axis data
    for idx, y_axis in enumerate(y_axes, start=2):
        # Create references to data
        x_values = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
        y_values = Reference(ws, min_col=idx, min_row=2, max_row=len(df)+1)
        
        series = Series(y_values, x_values, title_from_data=False)
        series.title = Reference(ws, min_col=idx, min_row=1, max_row=1)
        chart.series.append(series)
    
    # Add secondary Y-axis data if available
    if secondary_axes:
        # Create a secondary y-axis
        second_y_axis = chart.y_axis.copy()
        chart._axes.append(second_y_axis)
        second_y_axis.axId = 200  # arbitrary unique id
        second_y_axis.title = "Secondary Y-Axes"
        second_y_axis.crosses = "max"  # Place axis on the right
        
        # Add secondary series
        for idx, sec_axis in enumerate(secondary_axes, start=len(y_axes)+2):
            x_values = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
            y_values = Reference(ws, min_col=idx, min_row=2, max_row=len(df)+1)
            series = Series(y_values, x_values, title_from_data=False)
            series.title = Reference(ws, min_col=idx, min_row=1, max_row=1)
            chart.series.append(series)
            
            # Mark this series to use the secondary axis
            series.axId = second_y_axis.axId
    
    # Position the chart
    ws.add_chart(chart, "D5")
    logging.debug(f"Added line chart to worksheet")

def add_scatter_chart_to_worksheet(ws, df, x_axis, y_axes, secondary_axes):
    """Add a scatter chart to the worksheet
    
    Args:
        ws: The openpyxl worksheet object
        df: The pandas DataFrame containing source data
        x_axis: The name of the X-axis column
        y_axes: List of primary Y-axis column names
        secondary_axes: List of secondary Y-axis column names
    """
    chart = ScatterChart()
    chart.title = f"Scatter Chart: {', '.join(y_axes)} vs {x_axis}"
    chart.x_axis.title = x_axis
    chart.y_axis.title = "Y-Axes"
    
    # Get X values reference
    x_values = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
    
    # Add primary Y-axis series
    for idx, y_axis in enumerate(y_axes, start=2):
        y_values = Reference(ws, min_col=idx, min_row=2, max_row=len(df)+1)
        series = chart.series.append(y_values, x_values)
        series.title = Reference(ws, min_col=idx, min_row=1, max_row=1)
    
    # Add secondary Y-axis series if available
    if secondary_axes:
        # Create a secondary y-axis
        second_y_axis = chart.y_axis.copy()
        chart._axes.append(second_y_axis)
        second_y_axis.axId = 200  # arbitrary unique id
        second_y_axis.title = "Secondary Y-Axes"
        second_y_axis.crosses = "max"  # Place axis on the right
        
        # Add secondary series
        for idx, sec_axis in enumerate(secondary_axes, start=len(y_axes)+2):
            y_values = Reference(ws, min_col=idx, min_row=2, max_row=len(df)+1)
            series = chart.series.append(y_values, x_values)
            series.title = Reference(ws, min_col=idx, min_row=1, max_row=1)
            # Mark this series to use the secondary axis
            series.axId = second_y_axis.axId
    
    # Position the chart
    ws.add_chart(chart, "D5")
    logging.debug(f"Added scatter chart to worksheet")

def generate_excel_workbook(files, file_data_cache, x_axis, y_axes, secondary_axes, 
                          chart_type, scaling_options):
    """Generate an Excel workbook with charts for all files
    
    Args:
        files: List of file paths
        file_data_cache: Dictionary mapping file paths to DataFrames
        x_axis: Name of X-axis column
        y_axes: List of primary Y-axis column names
        secondary_axes: List of secondary Y-axis column names
        chart_type: Type of chart to generate ('line' or 'scatter')
        scaling_options: Dictionary of scaling options
    
    Returns:
        tuple: (Workbook, int, int) - Excel workbook, processed files count, skipped files count
    """
    from modules.data_processor import detect_header_row, read_csv_file, process_data_for_scaling, create_safe_sheet_name
    
    # Initialize workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    processed_files = 0
    skipped_files = 0
    
    for file_idx, file_path in enumerate(files):
        try:
            logging.info(f"Processing file {file_idx + 1}/{len(files)}: {os.path.basename(file_path)}")
            
            # Get data from cache or load if not present
            if file_path in file_data_cache:
                df = file_data_cache[file_path]
            else:
                header_row_idx = detect_header_row(file_path)
                df = read_csv_file(file_path, header_row_idx)
                file_data_cache[file_path] = df
            
            if df.empty:
                logging.warning(f"Skipping empty file: {os.path.basename(file_path)}")
                skipped_files += 1
                continue
                
            # Check if selected columns exist in the dataframe
            all_columns = [x_axis] + y_axes + secondary_axes
            missing_columns = [col for col in all_columns if col not in df.columns]
            
            if missing_columns:
                logging.warning(f"Skipping file {os.path.basename(file_path)} due to missing columns: {', '.join(missing_columns)}")
                skipped_files += 1
                continue
                
            # Process data for scaling
            processed_df = process_data_for_scaling(df, x_axis, y_axes, secondary_axes, scaling_options)
            
            # Create sheet name from filename
            sheet_name = create_safe_sheet_name(os.path.basename(file_path))
            if sheet_name in wb.sheetnames:
                sheet_name = f"{sheet_name}_{file_idx + 1}"
                
            # Create new worksheet
            ws = wb.create_sheet(title=sheet_name)
            
            # Populate data
            populate_worksheet(ws, processed_df, x_axis, y_axes, secondary_axes)
            
            # Add chart based on selected type
            if chart_type == 'scatter':
                add_scatter_chart_to_worksheet(ws, processed_df, x_axis, y_axes, secondary_axes)
            else:  # Default to line chart
                add_line_chart_to_worksheet(ws, processed_df, x_axis, y_axes, secondary_axes)
                
            processed_files += 1
            logging.info(f"Successfully processed file: {os.path.basename(file_path)}")
            
        except Exception as e:
            logging.error(f"Error processing file {os.path.basename(file_path)}: {str(e)}")
            skipped_files += 1
            
    # If all files were skipped and no sheets were created, add a dummy sheet
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="No Data")
        ws["A1"] = "No data could be processed from the selected files."
        
    logging.info(f"Excel workbook generation complete: {processed_files} files processed, {skipped_files} files skipped")
    return wb, processed_files, skipped_files

def export_figure(figure, output_filename, format_type):
    """Export a matplotlib figure to a file
    
    Args:
        figure: The matplotlib figure object
        output_filename: Path where to save the file
        format_type: Export format (png, pdf, svg, etc.)
    """
    try:
        # Ensure output directory exists
        output_dir = os.path.dirname(output_filename)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        # Save figure with high quality
        figure.savefig(
            output_filename, 
            format=format_type, 
            dpi=300, 
            bbox_inches='tight'
        )
        
        logging.info(f"Figure exported to {output_filename} in {format_type} format")
        return True
    except Exception as e:
        logging.error(f"Error exporting figure: {str(e)}")
        raise